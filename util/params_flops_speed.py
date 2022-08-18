import torch
import time
from fvcore.nn import FlopCountAnalysis, parameter_count
from model import get_model
from argparse import Namespace
import os
import openpyxl

import warnings
warnings.filterwarnings("ignore")


data_xlsx = []


def get_timepc():
	if torch.cuda.is_available():
		torch.cuda.synchronize()
	return time.perf_counter()


def eval_model(model_name, bs, img_size, gpu_id, speed=False):
	global data_xlsx
	print(f'===> {model_name}')
	x = torch.randn(bs, 3, img_size, img_size)
	model = Namespace()
	model.name = model_name
	model.model_kwargs = dict(pretrained=False, checkpoint_path='', ema=False, strict=True, num_classes=1000)
	net = get_model(model)
	net.eval()
	pre_cnt, cnt = 2, 5
	if gpu_id > -1:
		torch.cuda.set_device(gpu_id)
		x = x.cuda()
		net.cuda()
		pre_cnt, cnt = 5, 20
	flops = FlopCountAnalysis(net, x).total() / bs / 1e9
	params = parameter_count(net)[''] / 1e6
	if speed:
		with torch.no_grad():
			for _ in range(pre_cnt):
				y = net(x)
			t_s = get_timepc()
			for _ in range(cnt):
				y = net(x)
			t_e = get_timepc()
			flops = f'{flops:>6.3f}'
			params = f'{params:>6.3f}'
			speed = f'{bs * cnt / (t_e - t_s):>7.3f}'
			data_xlsx.append([model_name, params, flops, speed])
			ret_str = f'{model_name:>50}\t[GFLOPs: {flops}G]\t[Params: {params}M]\t[Speed: {speed}]\n'
	else:
		data_xlsx.append([model_name, params, flops])
		ret_str = f'{model_name:>50}\t[GFLOPs: {flops}G]\t[Params: {params}M]\n'
	return ret_str


def write_xlsx(file_path='flops_params_speed.xlsx', sheet_name='eatformer'):
	global data_xlsx
	if os.path.exists(file_path):
		workbook = openpyxl.load_workbook(file_path, read_only=False)
		if sheet_name in workbook.sheetnames:
			workbook.remove_sheet(workbook[sheet_name])
		sheet = workbook.create_sheet(sheet_name, index=None)
		# sheet = workbook.get_sheet_by_name(sheet_name)
	else:
		workbook = openpyxl.Workbook()
		sheet = workbook.active
		sheet.title = sheet_name
		# sheet = workbook.create_sheet(sheet_name, index=None)
	for i in range(len(data_xlsx)):
		for j in range(len(data_xlsx[i])):
			sheet.cell(row= i + 1, column= j + 1, value=str(data_xlsx[i][j]))
	workbook.save(file_path)

bs = 128
img_size = 224
gpu_id = 0  # -1 for CPU mode
speed = True
log_str = ''

# ============================== Classification ==============================
log_str += eval_model(model_name='timm_mobilenetv3_small_075', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
log_str += eval_model(model_name='eaformer_mobile', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)

# log_str += eval_model(model_name='timm_mobilenetv3_large_075', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='pvt_v2_b0', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='xcit_nano_12_p16', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='van_tiny', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='eaformer_lite', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)

# log_str += eval_model(model_name='timm_deit_tiny_patch16_224', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='eat_tiny_patch16_224', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_efficientnet_b0', bs=bs, img_size=224, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='coat_lite_tiny', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='ViTAE_basic_Tiny', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='ViTAE_basic_6M', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='xcit_tiny_12_p16', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='mpvit_tiny', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='eaformer_tiny', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='eaformer_tiny_384', bs=bs // 2, img_size=384, gpu_id=gpu_id, speed=speed)

# log_str += eval_model(model_name='timm_resnet18', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_efficientnet_b2', bs=bs, img_size=256, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='pvt_v2_b1', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='ViTAE_basic_13M', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='mpvit_xsmall', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='shunted_t', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='poolformer_s12', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='van_small', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='coat_lite_mini', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='xcit_tiny_24_p16', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='eaformer_mini', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)


# log_str += eval_model(model_name='timm_deit_small_patch16_224', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='eat_small_patch16_224', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_resnet50', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_regnetx_040', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_efficientnet_b4', bs=bs // 3, img_size=320, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_swin_tiny_patch4_window7_224', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='pvt_v2_b2', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='Twins_SVT_S', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='container_v1_light', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='uniformer_small', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='ViTAE_basic_Small', bs=bs // 2, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='ViTAE_basic_Small_384', bs=bs // 4, img_size=384, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='ViTAEv2_S', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='ViTAEv2_S_384', bs=bs // 4, img_size=384, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='DAT_T', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='mpvit_small', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='shunted_s', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='poolformer_s24', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='poolformer_s36', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='van_base', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='coat_lite_small', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='xcit_small_12_p8', bs=bs // 2, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='nat_tiny', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='crossformer_tiny', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='eaformer_small', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='eaformer_small_384', bs=bs // 4, img_size=384, gpu_id=gpu_id, speed=speed)

bs = 64
# log_str += eval_model(model_name='timm_resnet101', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_regnetx_080', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_efficientnet_b5', bs=bs // 4, img_size=456, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='pvt_v2_b3', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='xcit_small_24_p8', bs=bs // 2, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='eaformer_medium', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)

# log_str += eval_model(model_name='timm_vit_base_patch16_224', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_vit_base_patch16_224', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_vit_base_patch16_384', bs=bs // 2, img_size=384, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_deit_base_patch16_224', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='eat_base_patch16_224', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_resnet152', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_regnetx_160', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_efficientnet_b7', bs=bs // 16, img_size=600, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_swin_small_patch4_window7_224', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='timm_swin_base_patch4_window7_224', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='pvt_v2_b5', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='Twins_SVT_L', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='uniformer_base', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='DAT_B', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='mpvit_base', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='ViTAEv2_48M', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='shunted_b', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='poolformer_m48', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='van_large', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='coat_lite_medium', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='xcit_medium_24_p8', bs=bs // 4, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='nat_small', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='crossformer_base', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='eaformer_base', bs=bs, img_size=img_size, gpu_id=gpu_id, speed=speed)
# log_str += eval_model(model_name='eaformer_base_384', bs=bs // 4, img_size=384, gpu_id=gpu_id, speed=speed)

# write_xlsx('flops_params_speed.xlsx', 'sotas_GPU_V100')
# write_xlsx('flops_params_speed.xlsx', 'sotas_CPU')
print(log_str)
